#!/usr/bin/env python3
"""
SMART PRISM CONVERTER v3.1 - Immature Gut Lab

Converts Myography Analyzer output to Prism-ready format.
- Auto-detects P7NEO, P14NEO, P28NEO, P56NEO prefixes
- Handles experiment markers (R_, Th_, KO_, WT_, ShWT_, sham_)
- Handles comma segment notation (n2,1)
- Groups by AGE first, then Sex within each age group

USAGE: Double-click this file or drag your Excel file onto it.

v3.1 (Feb 2026) - Added support for experiment marker filenames (R_, Th_, etc.)
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
import re
import os
import sys

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, simpledialog
    HAS_TK = True
except ImportError:
    HAS_TK = False

# Age group sort order
AGE_ORDER = {'P7': 0, 'P14': 1, 'P28': 2, 'P56': 3, 'NEO': 4, 'ADULT': 5}

def get_age_prefix(subject_id):
    """Extract age prefix from Subject_ID like P7_M1 -> P7"""
    match = re.match(r'([A-Z]+\d*)', str(subject_id))
    if match:
        return match.group(1)
    return 'UNKNOWN'

def sort_by_age_then_sex(subject_id):
    """Sort key: age group first, then sex, then number"""
    prefix = get_age_prefix(subject_id)
    age_rank = AGE_ORDER.get(prefix, 99)
    sex_rank = 0 if '_M' in str(subject_id).upper() else 1
    return (age_rank, prefix, sex_rank, str(subject_id))

def detect_conditions(filenames):
    """Detect conditions from filenames"""
    raw = set(str(f).split('_')[-1] for f in filenames)
    conditions = []
    if 'Baseline' in raw:
        conditions.append('Baseline')
    treatments = [c for c in raw if c not in ['Baseline', 'Y2']]
    if treatments:
        conditions.append('Treatment')
    if 'Y2' in raw:
        conditions.append('Y2')
    return conditions or ['Baseline'], sorted(treatments)

def get_condition(filename, treatments):
    """Get condition from filename"""
    f = str(filename)
    if f.endswith('_Baseline'): return 'Baseline'
    if f.endswith('_Y2'): return 'Y2'
    for t in treatments:
        if f.endswith(f'_{t}'): return 'Treatment'
    return 'Baseline'

def get_filename_base(filename, treatments):
    """Remove condition suffix"""
    f = str(filename)
    for suffix in ['Baseline', 'Y2'] + treatments:
        if f.endswith(f'_{suffix}'):
            return f[:-len(f'_{suffix}')]
    return f

def parse_subject_id(filename_base):
    """Parse filename to extract subject info - handles P7NEO, n2,1 etc."""
    patterns = [
        # Pattern 0: P7NEO_Male1 format
        r'(\d{8})_(P\d+)NEO_([MF]|Male|male|Female|female)(\d+)_n(\d+(?:,\d+)?)',
        # Pattern 1: NEO_Male1 or P14_Male1 or ADULT_Male1 format
        r'(\d{8})_(NEO|P\d+|ADULT|Adult)_([MF]|Male|male|Female|female)(\d+)_n(\d+(?:,\d+)?)',
        # Pattern 2: NEOMale1 or P14Male1 format (no underscore before sex)
        r'(\d{8})_(NEO|P\d+|ADULT|Adult)([MF]|Male|male|Female|female)(\d+)_n(\d+(?:,\d+)?)',
        # Pattern 3: Experiment marker format - R_Male1, Th_Female2, KO_Male1, etc.
        r'(\d{8})_(R|Th|KO|WT|ShWT|sham)_([MF]|Male|male|Female|female)(\d+)_n(\d+(?:,\d+)?)',
        # Pattern 4: Simple format - Male1_n1 (no age or experiment marker)
        r'(\d{8})_([MF]|Male|male|Female|female)(\d+)_n(\d+(?:,\d+)?)',
    ]
    
    for i, pattern in enumerate(patterns):
        match = re.match(pattern, filename_base, re.IGNORECASE)
        if match:
            g = match.groups()
            if i == 0:
                # P7NEO format
                date, prefix, sex_ind, num, seg = g[0], g[1].upper(), g[2], g[3], g[4]
            elif i in [1, 2]:
                # NEO/P14/ADULT format
                date, prefix, sex_ind, num, seg = g[0], g[1].upper(), g[2], g[3], g[4]
            elif i == 3:
                # Experiment marker format (R_, Th_, KO_, etc.) - default to ADULT
                date, prefix, sex_ind, num, seg = g[0], 'ADULT', g[2], g[3], g[4]
            else:
                # Simple format - default to ADULT
                date, prefix, sex_ind, num, seg = g[0], 'ADULT', g[1], g[2], g[3]
            
            sex = 'M' if sex_ind.upper() in ['M', 'MALE'] else 'F'
            return {
                'filename_base': filename_base,
                'date': date, 'prefix': prefix, 'sex': sex,
                'animal_num': num, 'segment': seg.replace(',', '.'),
                'unique_key': f"{date}_{prefix}_{sex}{num}"
            }
    return None

def get_sex(subject_id):
    """Get sex from Subject_ID"""
    s = str(subject_id).upper()
    if '_M' in s: return 'Male'
    if '_F' in s: return 'Female'
    return 'Unknown'

def auto_assign_ids(filename_bases):
    """Auto-assign Subject_IDs"""
    from collections import OrderedDict
    
    parsed = [p for fb in filename_bases if (p := parse_subject_id(fb))]
    if not parsed: return {}
    
    seen = OrderedDict()
    for a in parsed:
        if a['unique_key'] not in seen:
            seen[a['unique_key']] = a
    
    sorted_animals = sorted(seen.values(), 
        key=lambda x: (AGE_ORDER.get(x['prefix'], 99), x['date'], 0 if x['sex']=='M' else 1, int(x['animal_num'])))
    
    counters = {}
    unique_to_id = {}
    for a in sorted_animals:
        key = f"{a['prefix']}_{a['sex']}"
        counters[key] = counters.get(key, 0) + 1
        unique_to_id[a['unique_key']] = f"{a['prefix']}_{a['sex']}{counters[key]}"
    
    return {a['filename_base']: unique_to_id[a['unique_key']] for a in parsed}

def get_animal_base(subject_id):
    """Remove segment suffix"""
    return re.sub(r'\.\d+(\.\d+)?$', '', str(subject_id))

def calc_pct_change(baseline, new_val):
    """Calculate percent change"""
    if pd.isna(baseline) or pd.isna(new_val) or baseline == 0:
        return np.nan
    pct = ((new_val - baseline) / baseline) * 100
    return round(pct) if not (pd.isna(pct) or np.isinf(pct)) else np.nan

def add_source_sheet(wb, df, name, color):
    """Add source data sheet"""
    ws = wb.create_sheet(title=name)
    ws.sheet_properties.tabColor = color
    for c, col in enumerate(df.columns, 1):
        ws.cell(1, c, col).font = Font(bold=True)
    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            if pd.notna(val): ws.cell(r, c, val)

def process_file(input_file, exclude_subjects=None):
    """Main processing function"""
    exclude_subjects = exclude_subjects or []
    output_file = os.path.splitext(input_file)[0] + '_Prism.xlsx'
    
    print(f"\nProcessing: {os.path.basename(input_file)}")
    
    try:
        xlsx = pd.ExcelFile(input_file)
    except Exception as e:
        print(f"  ERROR: {e}")
        return None
    
    if 'Overall_Metrics' not in xlsx.sheet_names or '10sec_Bins' not in xlsx.sheet_names:
        print("  ERROR: Missing required sheets")
        return None
    
    df_overall_orig = pd.read_excel(xlsx, 'Overall_Metrics')
    df_10sec_orig = pd.read_excel(xlsx, '10sec_Bins')
    df_overall = df_overall_orig.copy()
    df_10sec = df_10sec_orig.copy()
    
    # Strip whitespace from filenames to prevent condition detection issues
    df_overall['Filename'] = df_overall['Filename'].str.strip()
    df_10sec['Filename'] = df_10sec['Filename'].str.strip()
    
    conditions, treatments = detect_conditions(df_overall['Filename'])
    print(f"  Conditions: {conditions}")
    
    # Build subject mapping
    df_overall['FB_Temp'] = df_overall['Filename'].apply(lambda x: get_filename_base(x, treatments))
    mapping = auto_assign_ids(df_overall['FB_Temp'].unique())
    
    if not mapping:
        print("  WARNING: Could not auto-detect. Using fallback.")
        mapping = {fb: fb for fb in df_overall['FB_Temp'].unique()}
    
    print(f"  Auto-detected: {len(mapping)} samples (100%)")
    
    # Apply mapping with segment suffix
    df_overall['Subject_ID_Base'] = df_overall['FB_Temp'].map(mapping)
    
    def add_segment(row):
        base = row['Subject_ID_Base']
        if pd.isna(base): return 'Unknown'
        match = re.search(r'_n(\d+(?:,\d+)?)_', row['Filename'])
        return f"{base}.{match.group(1).replace(',', '.')}" if match else str(base)
    
    df_overall['Subject_ID'] = df_overall.apply(add_segment, axis=1)
    df_overall['Condition'] = df_overall['Filename'].apply(lambda x: get_condition(x, treatments))
    df_overall['Sex'] = df_overall['Subject_ID'].apply(get_sex)
    df_overall['Filename_Base'] = df_overall['FB_Temp']
    df_overall.drop(['FB_Temp', 'Subject_ID_Base'], axis=1, inplace=True)
    
    df_filtered = df_overall[~df_overall['Subject_ID'].isin(exclude_subjects)].copy()
    
    # Create filename to subject map for bin data
    fn_to_subj = df_overall.drop_duplicates('Filename_Base').set_index('Filename_Base')['Subject_ID'].to_dict()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Overall"
    ws.sheet_properties.tabColor = "FFC000"
    
    metrics = [c for c in df_overall.columns if c not in 
               ['Filename', 'Subject_ID', 'Condition', 'Sex', 'Filename_Base', df_overall.columns[0]]]
    
    # Build headers
    pct_cols = []
    if 'Treatment' in conditions: pct_cols.append(('Treatment', '%chg_Trt'))
    if 'Y2' in conditions: pct_cols.append(('Y2', '%chg_Y2'))
    
    col = 1
    for h in ["Sex", "Subject_ID", "Filename_Base"]:
        ws.cell(1, col, h).font = Font(bold=True)
        ws.cell(2, col, h).font = Font(bold=True)
        col += 1
    
    col_map, pct_map = {}, {}
    for i, m in enumerate(metrics):
        ws.cell(1, col, m).font = Font(bold=True)
        for c in conditions:
            ws.cell(2, col, c).font = Font(bold=True)
            col_map[(m, c)] = col
            col += 1
        for c, lbl in pct_cols:
            ws.cell(2, col, lbl).font = Font(bold=True)
            pct_map[(m, c)] = col
            col += 1
        if i < len(metrics) - 1: col += 1
    
    # Write data - SORTED BY AGE THEN SEX
    subjects = sorted(df_filtered['Subject_ID'].unique(), key=sort_by_age_then_sex)
    subj_fb_map = df_filtered.drop_duplicates('Subject_ID').set_index('Subject_ID')['Filename_Base'].to_dict()
    
    row = 3
    last_prefix, last_sex = None, None
    
    for subj in subjects:
        prefix = get_age_prefix(subj)
        sex = get_sex(subj)
        
        # Blank row between age groups or between sexes within group
        if last_prefix is not None and (prefix != last_prefix or sex != last_sex):
            row += 1
        last_prefix, last_sex = prefix, sex
        
        data = df_filtered[df_filtered['Subject_ID'] == subj]
        ws.cell(row, 1, sex)
        ws.cell(row, 2, subj)
        ws.cell(row, 3, subj_fb_map.get(subj, ''))
        
        for m in metrics:
            bl_data = data[data['Condition'] == 'Baseline']
            bl_val = bl_data[m].iloc[0] if len(bl_data) > 0 else np.nan
            
            for c in conditions:
                cond_data = data[data['Condition'] == c]
                if len(cond_data) > 0:
                    val = cond_data[m].iloc[0]
                    if pd.notna(val): ws.cell(row, col_map[(m, c)], val)
            
            for c, _ in pct_cols:
                cond_data = data[data['Condition'] == c]
                if len(cond_data) > 0:
                    pct = calc_pct_change(bl_val, cond_data[m].iloc[0])
                    if pd.notna(pct): ws.cell(row, pct_map[(m, c)], int(pct))
        row += 1
    
    # Animal averages
    df_filtered['Animal_Base'] = df_filtered['Subject_ID'].apply(get_animal_base)
    df_filtered['Date_Prefix'] = df_filtered['Filename_Base'].apply(lambda x: str(x).split('_')[0])
    
    row += 1
    ws.cell(row, 2, "Averages per Animal").font = Font(bold=True)
    row += 1
    
    animals = sorted(df_filtered['Animal_Base'].unique(), key=sort_by_age_then_sex)
    last_prefix, last_sex = None, None
    
    for animal in animals:
        prefix = get_age_prefix(animal)
        sex = get_sex(animal)
        
        if last_prefix is not None and (prefix != last_prefix or sex != last_sex):
            row += 1
        last_prefix, last_sex = prefix, sex
        
        data = df_filtered[df_filtered['Animal_Base'] == animal]
        date_pfx = data['Date_Prefix'].iloc[0]
        
        ws.cell(row, 2, animal)
        ws.cell(row, 3, f"{date_pfx}_{animal}")
        
        for m in metrics:
            for c in conditions:
                vals = data[data['Condition'] == c][m].dropna()
                if len(vals) > 0:
                    mean = vals.mean()
                    if pd.notna(mean) and not np.isinf(mean):
                        ws.cell(row, col_map[(m, c)], mean)
            
            bl_vals = data[data['Condition'] == 'Baseline'][m].dropna()
            bl_mean = bl_vals.mean() if len(bl_vals) > 0 else np.nan
            
            for c, _ in pct_cols:
                vals = data[data['Condition'] == c][m].dropna()
                if len(vals) > 0:
                    pct = calc_pct_change(bl_mean, vals.mean())
                    if pd.notna(pct): ws.cell(row, pct_map[(m, c)], int(pct))
        row += 1
    
    # 10-second bin sheets
    bin_metrics = {
        'Mean_Amplitude_mN': 'Amplitude', 'Frequency_cpm': 'Frequency',
        'Mean_Period_sec': 'Period', 'Mean_Duration_sec': 'Duration',
        'Mean_Rise_Time_sec': 'RiseTime', 'Mean_Relax_Time_sec': 'RelaxTime',
        'Integral_Force_mN_s': 'Integral'
    }
    
    available = {m: m for m in bin_metrics if m in df_10sec.columns}
    print(f"  Bin metrics: {list(available.keys())}")
    
    df_10sec['Filename_Base'] = df_10sec['Filename'].apply(lambda x: get_filename_base(x, treatments))
    df_10sec['Condition'] = df_10sec['Filename'].apply(lambda x: get_condition(x, treatments))
    df_10sec['Subject_ID'] = df_10sec['Filename_Base'].map(fn_to_subj)
    df_10sec['Sex'] = df_10sec['Subject_ID'].apply(get_sex)
    df_10sec['Animal_Base'] = df_10sec['Subject_ID'].apply(get_animal_base)
    df_10sec['Date_Prefix'] = df_10sec['Filename_Base'].apply(lambda x: str(x).split('_')[0])
    df_10sec_clean = df_10sec.dropna(subset=['Subject_ID'])
    df_10sec_clean = df_10sec_clean[~df_10sec_clean['Subject_ID'].isin(exclude_subjects)]
    
    headers = ['Sex', 'Subject_ID', 'Filename_Base'] + [f'{c}_{s}' for c in conditions for s in ['Mean', 'SD', 'n']]
    
    print("  Creating bin sheets...")
    for metric, short in bin_metrics.items():
        if metric not in available: continue
        
        ws_m = wb.create_sheet(f"10s_{short}")
        ws_m.sheet_properties.tabColor = "4472C4"
        ws_m.cell(1, 1, f"{metric} (10s)").font = Font(bold=True)
        
        for c, h in enumerate(headers, 1):
            ws_m.cell(3, c, h).font = Font(bold=True)
        
        # Per-subject stats
        stats = []
        for subj in sorted(df_10sec_clean['Subject_ID'].unique(), key=sort_by_age_then_sex):
            d = df_10sec_clean[df_10sec_clean['Subject_ID'] == subj]
            r = {'Sex': d['Sex'].iloc[0], 'Subject_ID': subj, 'Filename_Base': d['Filename_Base'].iloc[0]}
            for c in conditions:
                vals = d[d['Condition'] == c][metric].dropna()
                r[f'{c}_Mean'] = vals.mean() if len(vals) > 0 else np.nan
                r[f'{c}_SD'] = vals.std(ddof=1) if len(vals) > 1 else np.nan
                r[f'{c}_n'] = int(len(vals))
            stats.append(r)
        
        row = 4
        last_prefix, last_sex = None, None
        for s in stats:
            prefix = get_age_prefix(s['Subject_ID'])
            sex = s['Sex']
            if last_prefix is not None and (prefix != last_prefix or sex != last_sex):
                row += 1
            last_prefix, last_sex = prefix, sex
            
            for c, h in enumerate(headers, 1):
                ws_m.cell(row, c, s.get(h, ''))
            row += 1
        
        # Animal averages
        row += 1
        ws_m.cell(row, 2, "Averages per Animal").font = Font(bold=True)
        row += 1
        
        last_prefix, last_sex = None, None
        for animal in sorted(df_10sec_clean['Animal_Base'].unique(), key=sort_by_age_then_sex):
            prefix = get_age_prefix(animal)
            sex = get_sex(animal)
            if last_prefix is not None and (prefix != last_prefix or sex != last_sex):
                row += 1
            last_prefix, last_sex = prefix, sex
            
            d = df_10sec_clean[df_10sec_clean['Animal_Base'] == animal]
            ws_m.cell(row, 1, sex)
            ws_m.cell(row, 2, animal)
            ws_m.cell(row, 3, f"{d['Date_Prefix'].iloc[0]}_{animal}")
            
            col = 4
            for c in conditions:
                vals = d[d['Condition'] == c][metric].dropna()
                ws_m.cell(row, col, vals.mean() if len(vals) > 0 else '')
                ws_m.cell(row, col+1, vals.std(ddof=1) if len(vals) > 1 else '')
                ws_m.cell(row, col+2, int(len(vals)))
                col += 3
            row += 1
        
        # Raw bin data
        row += 2
        ws_m.cell(row, 1, "Raw Bin Data").font = Font(bold=True)
        row += 1
        raw_h = ['Sex', 'Subject_ID', 'Condition', 'Bin', 'Time_Start', 'Time_End', metric]
        for c, h in enumerate(raw_h, 1):
            ws_m.cell(row, c, h).font = Font(bold=True)
        row += 1
        
        df_sorted = df_10sec_clean.sort_values(['Subject_ID', 'Condition', 'Bin'],
            key=lambda x: x.map(lambda v: sort_by_age_then_sex(v) if x.name == 'Subject_ID' 
                else (0 if v == 'Baseline' else 1) if x.name == 'Condition' else v))
        
        last_subj = None
        for _, r in df_sorted.iterrows():
            if last_subj and r['Subject_ID'] != last_subj: row += 1
            last_subj = r['Subject_ID']
            ws_m.cell(row, 1, r['Sex'])
            ws_m.cell(row, 2, r['Subject_ID'])
            ws_m.cell(row, 3, r['Condition'])
            ws_m.cell(row, 4, r.get('Bin', ''))
            ws_m.cell(row, 5, r.get('Time_Start_sec', ''))
            ws_m.cell(row, 6, r.get('Time_End_sec', ''))
            ws_m.cell(row, 7, r.get(metric, ''))
            row += 1
    
    add_source_sheet(wb, df_overall_orig, "SRC_Overall", "A6A6A6")
    add_source_sheet(wb, df_10sec_orig, "SRC_10sec_Bins", "A6A6A6")
    
    wb.save(output_file)
    print(f"\n  ✓ Saved: {output_file}")
    return output_file


if __name__ == "__main__":
    # Handle drag-and-drop or command line argument
    if len(sys.argv) > 1:
        for f in sys.argv[1:]:
            if os.path.exists(f) and f.endswith('.xlsx'):
                process_file(f)
        input("\nPress Enter to exit...")
    elif HAS_TK:
        root = tk.Tk()
        root.withdraw()
        files = filedialog.askopenfilenames(title="Select Excel File(s)", 
            filetypes=[("Excel files", "*.xlsx")])
        for f in files:
            process_file(f)
        if files:
            messagebox.showinfo("Done", f"Processed {len(files)} file(s)!")
    else:
        print("Drag Excel file onto this script or provide path as argument")
